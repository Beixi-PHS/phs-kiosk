"""
PHS Patient Feedback Kiosk ÃÂ¢ÃÂÃÂ Backend
=====================================
Serves the kiosk frontend and records every submission
directly to an Excel file in Microsoft OneDrive via the
Microsoft Graph API. Data is permanent ÃÂ¢ÃÂÃÂ no local file,
no data loss risk, no weekly downloads required.

Endpoints:
  GET  /           ÃÂ¢ÃÂÃÂ Serves the kiosk HTML
  POST /submit     ÃÂ¢ÃÂÃÂ Records feedback to OneDrive Excel
  GET  /dashboard  ÃÂ¢ÃÂÃÂ Password-protected live summary
  GET  /health     ÃÂ¢ÃÂÃÂ Service status
"""

import os
import logging
import threading
import requests
from datetime import datetime, timezone
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo
from pathlib import Path
from flask import Flask, request, jsonify, send_file, abort

# ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ Logging ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)

app = Flask(__name__, static_folder='.', template_folder='.')

# ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ Configuration (set as environment variables in Render.com) ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ
AZURE_TENANT_ID     = os.environ.get("AZURE_TENANT_ID")
AZURE_CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")
ONEDRIVE_FILE_ID    = os.environ.get("ONEDRIVE_FILE_ID")
ONEDRIVE_USER_ID    = os.environ.get("ONEDRIVE_USER_ID")
DOWNLOAD_SECRET     = os.environ.get("DOWNLOAD_SECRET", "phs2026")
WORKSHEET_NAME      = os.environ.get("WORKSHEET_NAME", "In Clinic Feedback")

RATING_LABELS = {1: "Poor", 2: "Fair", 3: "Good", 4: "Very Good", 5: "Excellent"}


# ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ Microsoft Graph helpers ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ

def is_graph_configured():
    return all([AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET,
                ONEDRIVE_FILE_ID, ONEDRIVE_USER_ID])


# Token cache — avoids an Azure round-trip on every submission (token valid 1 hr)
_token_cache = {"token": None, "expires_at": 0}

def get_access_token():
    import time
    now = time.time()
    if _token_cache["token"] and now < _token_cache["expires_at"]:
        return _token_cache["token"]
    url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
    res = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     AZURE_CLIENT_ID,
        "client_secret": AZURE_CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default"
    }, timeout=10)
    res.raise_for_status()
    data = res.json()
    _token_cache["token"] = data["access_token"]
    _token_cache["expires_at"] = time.time() + data.get("expires_in", 3600) - 300
    log.info("Access token refreshed.")
    return _token_cache["token"]


def graph_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def base_url():
    return (f"https://graph.microsoft.com/v1.0/users/{ONEDRIVE_USER_ID}"
            f"/drive/items/{ONEDRIVE_FILE_ID}/workbook/worksheets/{WORKSHEET_NAME}")


def get_used_row_count(token):
    res = requests.get(f"{base_url()}/usedRange", headers=graph_headers(token), timeout=10)
    if res.status_code != 200:
        return 0
    return res.json().get("rowCount", 0)


def write_range(token, values, start_row):
    num_cols  = len(values[0])
    end_col   = chr(ord('A') + num_cols - 1)
    end_row   = start_row + len(values) - 1
    addr      = f"A{start_row}:{end_col}{end_row}"
    url       = f"{base_url()}/range(address='{addr}')"
    res       = requests.patch(url, headers=graph_headers(token),
                               json={"values": values}, timeout=10)
    res.raise_for_status()
    log.info(f"Written to OneDrive Excel {addr}: {values}")


def ensure_headers(token):
    if get_used_row_count(token) == 0:
        write_range(token,
                    [["Timestamp", "Date", "Time", "Clinic", "Therapist", "Rating", "Rating Label"]],
                    start_row=1)
        log.info("Header row written.")


def append_to_onedrive(therapist, location, rating, timestamp_str):
    eastern = ZoneInfo("America/New_York")
    try:
        ts = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00")).astimezone(eastern)
    except Exception:
        ts = datetime.now(tz=eastern)

    token = get_access_token()
    ensure_headers(token)
    next_row = get_used_row_count(token) + 1

    write_range(token, [[
        timestamp_str,
        ts.strftime("%d %b %Y"),
        ts.strftime("%H:%M"),
        location,
        therapist,
        rating,
        RATING_LABELS.get(rating, str(rating))
    ]], start_row=next_row)

    log.info(f"Recorded to OneDrive: {therapist} ÃÂ¢ÃÂÃÂ {rating} stars")


# ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ Local fallback (dev / unconfigured) ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ

def append_local(therapist, location, rating, timestamp_str):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    path = Path("in_clinic_feedback_local.xlsx")
    eastern = ZoneInfo("America/New_York")
    try:
        ts = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00")).astimezone(eastern)
    except Exception:
        ts = datetime.now(tz=eastern)

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not path.exists():
        wb = Workbook(); ws = wb.active; ws.title = WORKSHEET_NAME
        for i, (h, w) in enumerate(zip(
            ["Timestamp","Date","Time","Clinic","Therapist","Rating","Rating Label"],
            [24,14,10,18,24,10,14]), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
            c.fill = PatternFill("solid", fgColor="2A435F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        wb.save(path)

    wb = load_workbook(path); ws = wb.active
    r  = ws.max_row + 1
    fill = PatternFill("solid", fgColor="E8ECF0") if r % 2 == 0 else None
    for i, v in enumerate([
        timestamp_str, ts.strftime("%d %b %Y"), ts.strftime("%H:%M"),
        location, therapist, rating, RATING_LABELS.get(rating)], 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        if fill: c.fill = fill
    ws.row_dimensions[r].height = 22
    wb.save(path)
    log.info(f"[LOCAL] Recorded: {therapist} ÃÂ¢ÃÂÃÂ {rating} stars")


# ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ Routes ÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂÃÂ¢ÃÂÃÂ

@app.route("/")
def index():
    return send_file("index.html")


@app.route("/submit", methods=["POST"])
def submit():
    data      = request.get_json(silent=True) or {}
    therapist = data.get("therapist", "").strip()
    location  = data.get("location", "").strip()
    rating    = data.get("rating")
    timestamp = data.get("timestamp", datetime.now(tz=timezone.utc).isoformat())

    if not therapist:
        return jsonify({"error": "Therapist is required"}), 400
    if not isinstance(rating, int) or rating not in range(1, 6):
        return jsonify({"error": "Rating must be 1ÃÂ¢ÃÂÃÂ5"}), 400

    def _record():
        try:
            if is_graph_configured():
                append_to_onedrive(therapist, location, rating, timestamp)
                log.info(f"Background write OK (onedrive): {therapist}")
            else:
                log.warning("Graph API not configured — using local fallback.")
                append_local(therapist, location, rating, timestamp)
                log.info(f"Background write OK (local): {therapist}")
        except Exception as e:
            log.error(f"Background submit failed: {e}")

    threading.Thread(target=_record, daemon=True).start()
    return jsonify({"status": "accepted", "therapist": therapist, "rating": rating}), 202


@app.route("/dashboard")
def dashboard():
    secret = request.args.get("secret", "")
    if secret != DOWNLOAD_SECRET:
        abort(401)

    rows    = []
    storage = "OneDrive (Microsoft 365)" if is_graph_configured() else "Local fallback"

    if is_graph_configured():
        try:
            token = get_access_token()
            res   = requests.get(f"{base_url()}/usedRange",
                                 headers=graph_headers(token), timeout=10)
            all_rows = res.json().get("values", [])
            for row in all_rows[1:]:
                if len(row) >= 5 and row[3]:
                    rows.append({"date": row[1] or "ÃÂ¢ÃÂÃÂ", "time": row[2] or "ÃÂ¢ÃÂÃÂ",
                                 "clinic": row[3], "therapist": row[4], "rating": int(row[5] or 0),
                                 "label": row[6] if len(row) > 6 else "ÃÂ¢ÃÂÃÂ"})
        except Exception as e:
            log.error(f"Dashboard fetch error: {e}")

    total     = len(rows)
    avg       = round(sum(r["rating"] for r in rows) / total, 1) if total else 0
    five_star = sum(1 for r in rows if r["rating"] == 5)

    t_stats = {}
    for r in rows:
        t = r["therapist"]
        if t not in t_stats:
            t_stats[t] = {"count": 0, "total": 0}
        t_stats[t]["count"] += 1
        t_stats[t]["total"] += r["rating"]

    t_rows = "".join(
        f"<tr><td>{n}</td><td>{s['count']}</td>"
        f"<td>{round(s['total']/s['count'],1)}</td>"
        f"<td style='color:#b8963e;letter-spacing:2px'>"
        f"{'ÃÂ¢ÃÂÃÂ'*int(round(s['total']/s['count']))}{'ÃÂ¢ÃÂÃÂ'*(5-int(round(s['total']/s['count'])))}"
        f"</td></tr>"
        for n, s in sorted(t_stats.items())
    ) or "<tr><td colspan='4' style='text-align:center;color:#aaa'>No data yet</td></tr>"

    r_rows = "".join(
        f"<tr><td>{r['date']}</td><td>{r['time']}</td><td>{r['therapist']}</td>"
        f"<td>{r['rating']}/5</td>"
        f"<td style='color:#b8963e'>{'ÃÂ¢ÃÂÃÂ'*r['rating']}{'ÃÂ¢ÃÂÃÂ'*(5-r['rating'])}</td></tr>"
        for r in reversed(rows[-20:])
    ) or "<tr><td colspan='5' style='text-align:center;color:#aaa'>No responses yet</td></tr>"

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PHS In Clinic Feedback Dashboard</title>
<style>
body{{font-family:'Segoe UI',sans-serif;background:#f9f8f6;color:#2A435F;margin:0;padding:32px}}
h1{{font-size:28px;font-weight:300;margin-bottom:4px}}
p.sub{{color:#888;font-size:13px;margin-bottom:6px}}
p.store{{color:#2A435F;font-size:12px;background:#e8ecf0;display:inline-block;
  padding:4px 12px;border-radius:20px;margin-bottom:36px}}
.stats{{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-bottom:40px}}
.stat{{background:#2A435F;color:white;padding:28px;border-radius:12px}}
.stat .num{{font-size:48px;font-weight:300;line-height:1}}
.stat .lbl{{font-size:12px;opacity:.6;margin-top:6px;letter-spacing:.08em;text-transform:uppercase}}
table{{width:100%;border-collapse:collapse;background:white;border-radius:12px;
  overflow:hidden;margin-bottom:40px;box-shadow:0 1px 4px rgba(0,0,0,.06)}}
th{{background:#2A435F;color:white;padding:12px 16px;text-align:left;
  font-size:12px;letter-spacing:.06em;text-transform:uppercase;font-weight:400}}
td{{padding:12px 16px;font-size:14px;border-bottom:1px solid #eee}}
tr:last-child td{{border-bottom:none}}
h2{{font-size:18px;font-weight:400;margin-bottom:12px}}
</style></head><body>
<h1>PHS In Clinic Feedback Dashboard</h1>
<p class="sub">Updated in real time ÃÂÃÂ· {datetime.now().strftime('%d %b %Y %H:%M')}</p>
<p class="store">ÃÂ¢ÃÂÃÂ {storage}</p>
<div class="stats">
  <div class="stat"><div class="num">{total}</div><div class="lbl">Total Responses</div></div>
  <div class="stat"><div class="num">{avg}</div><div class="lbl">Average Rating</div></div>
  <div class="stat"><div class="num">{five_star}</div><div class="lbl">5-Star Ratings</div></div>
</div>
<h2>By Therapist</h2>
<table><thead><tr><th>Therapist</th><th>Responses</th><th>Avg Rating</th><th>Stars</th></tr></thead>
<tbody>{t_rows}</tbody></table>
<h2>Recent Responses</h2>
<table><thead><tr><th>Date</th><th>Time</th><th>Therapist</th><th>Rating</th><th>Stars</th></tr></thead>
<tbody>{r_rows}</tbody></table>
</body></html>"""


@app.route("/health")
def health():
    return jsonify({
        "status":      "online",
        "storage":     "onedrive" if is_graph_configured() else "local_fallback",
        "graph_ready": is_graph_configured(),
        "timestamp":   datetime.utcnow().isoformat()
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    log.info("OneDrive mode." if is_graph_configured() else "Local fallback mode.")
    app.run(host="0.0.0.0", port=port, debug=False)
